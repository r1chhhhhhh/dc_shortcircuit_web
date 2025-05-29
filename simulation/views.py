from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, JsonResponse, Http404
from django.contrib import messages
from django.core.files.base import ContentFile
from django.conf import settings
import os
import json
import base64
from io import BytesIO
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend
import matplotlib.pyplot as plt
from dataclasses import dataclass
from .models import SimulationRun
from .forms import SimulationForm
import uuid
import time
import matplotlib
matplotlib.use('Agg')  # 在最开始就设置后端
import matplotlib.pyplot as plt


@dataclass
class SimulationInputs:
    sourceVoltage: float = 750.0
    sourceInternalResistance: float = 0.01
    lineResistance: float = 0.02
    faultResistance: float = 0.01
    faultDistance: float = 1.0
    numSubstations: int = 3
    substationSpacing: float = 5.0
    protectionTripTime: float = 0.1
    protectionTripCurrent: float = 4000
    systemInductance: float = 0.001
    lineInductance: float = 0.0005
    simulationTimeStep: float = 0.1
    simulationDuration: float = 350.0
    lineLength: float = 10.0


@dataclass
class SimulationResults:
    totalFaultCurrent: float = 0.0
    energyReleased: float = 0.0
    thermalStress: float = 0.0
    protectionWillTrip: bool = False
    tripTime: float = 0.0
    shortestDistanceCurrent: float = 0.0
    longestDistanceCurrent: float = 0.0
    shortestDistanceDiDt: float = 0.0
    longestDistanceDiDt: float = 0.0
    peakDiDt: float = 0.0
    steadyStateDiDt: float = 0.0


class TimeSeriesData:
    def __init__(self):
        self.timePoints = []
        self.current = []
        self.diDt = []
        self.filteredDiDt = []
        self.voltage = []
        self.timePointCount = 0
        self.peakDiDt = 0.0
        self.peakCurrent = 0.0


class DCShortCircuitCalculation:
    """Core calculation engine separated from GUI"""
    
    def __init__(self, inputs):
        self.inputs = inputs
        self.results = SimulationResults()
        self.time_series_data = TimeSeriesData()
        self.substation_locations = []
        self.fault_currents = []
        self.voltage_drops = []

    def determine_contributing_substations(self, fault_location, line_length, num_substations):
        """Determine which substations contribute current to the fault"""
        contributing_substations = []
        for i in range(num_substations):
            if num_substations > 1:
                substation_location = i * (line_length / (num_substations - 1))
            else:
                substation_location = 0
            contributing_substations.append(substation_location)
        return contributing_substations

    def calculate_short_circuit(self):
        """Calculate short circuit parameters and time domain analysis"""
        fault_location = self.inputs.faultDistance
        contributing_substations = self.determine_contributing_substations(
            fault_location, self.inputs.lineLength, self.inputs.numSubstations
        )

        total_fault_current = 0
        fault_currents = []
        voltage_drops = []

        for substation_location in contributing_substations:
            distance_to_fault = abs(fault_location - substation_location)
            resistance_to_fault = (self.inputs.sourceInternalResistance +
                                   (distance_to_fault * self.inputs.lineResistance) +
                                   self.inputs.faultResistance)
            inductance_to_fault = (self.inputs.systemInductance +
                                   (distance_to_fault * self.inputs.lineInductance))

            if resistance_to_fault <= 0:
                resistance_to_fault = 1e-7

            fault_current = self.inputs.sourceVoltage / resistance_to_fault
            voltage_drop = fault_current * (distance_to_fault * self.inputs.lineResistance)

            fault_currents.append(fault_current)
            voltage_drops.append(voltage_drop)
            total_fault_current += fault_current

        # Calculate results
        energy_released = (total_fault_current ** 2) * self.inputs.faultResistance * self.inputs.protectionTripTime
        thermal_stress = total_fault_current * np.sqrt(self.inputs.protectionTripTime)
        protection_will_trip = (total_fault_current >= self.inputs.protectionTripCurrent)

        if protection_will_trip:
            over_current_ratio = total_fault_current / self.inputs.protectionTripCurrent
            if over_current_ratio > 1:
                trip_time = self.inputs.protectionTripTime / (over_current_ratio - 0.9)
                if trip_time < 0.05:
                    trip_time = 0.05
            else:
                trip_time = 100
        else:
            trip_time = 100

        # Calculate shortest and longest distance currents
        shortest_resistance = self.inputs.sourceInternalResistance + self.inputs.faultResistance
        if shortest_resistance <= 0:
            shortest_resistance = 1e-7
        shortest_distance_current = self.inputs.sourceVoltage / shortest_resistance

        if self.inputs.numSubstations > 1:
            longest_distance = self.inputs.substationSpacing / 2
        else:
            longest_distance = self.inputs.lineLength

        longest_resistance = (self.inputs.sourceInternalResistance +
                              (longest_distance * self.inputs.lineResistance) +
                              self.inputs.faultResistance)
        if longest_resistance <= 0:
            longest_resistance = 1e-7
        longest_distance_current = self.inputs.sourceVoltage / longest_resistance

        # Calculate di/dt values
        shortest_inductance = self.inputs.systemInductance
        if shortest_inductance <= 0:
            shortest_inductance = 1e-7

        longest_inductance = self.inputs.systemInductance + (longest_distance * self.inputs.lineInductance)
        if longest_inductance <= 0:
            longest_inductance = 1e-7

        shortest_distance_didt = self.inputs.sourceVoltage / shortest_inductance
        longest_distance_didt = self.inputs.sourceVoltage / longest_inductance

        # Store results
        self.results.totalFaultCurrent = total_fault_current
        self.results.energyReleased = energy_released
        self.results.thermalStress = thermal_stress
        self.results.protectionWillTrip = protection_will_trip
        self.results.tripTime = trip_time
        self.results.shortestDistanceCurrent = shortest_distance_current
        self.results.longestDistanceCurrent = longest_distance_current
        self.results.shortestDistanceDiDt = shortest_distance_didt
        self.results.longestDistanceDiDt = longest_distance_didt

        # Store substation data
        self.substation_locations = contributing_substations
        self.fault_currents = fault_currents
        self.voltage_drops = voltage_drops

        # Perform time domain analysis
        self.perform_time_domain_analysis(fault_location)

    def perform_time_domain_analysis(self, fault_location):
        """Perform time domain analysis of fault current"""
        num_points = int(self.inputs.simulationDuration / self.inputs.simulationTimeStep) + 1
        
        self.time_series_data = TimeSeriesData()
        self.time_series_data.timePointCount = num_points

        distance_to_fault = fault_location
        resistance_to_fault = (self.inputs.sourceInternalResistance +
                               (distance_to_fault * self.inputs.lineResistance) +
                               self.inputs.faultResistance)
        inductance_to_fault = (self.inputs.systemInductance +
                               (distance_to_fault * self.inputs.lineInductance))

        time_const = inductance_to_fault / resistance_to_fault
        steady_state_current = self.inputs.sourceVoltage / resistance_to_fault
        dt = self.inputs.simulationTimeStep / 1000
        prev_current = 0

        for i in range(num_points):
            t = i * dt * 1000
            self.time_series_data.timePoints.append(t)

            if t < 0:
                current_value = 0
            else:
                current_value = steady_state_current * (1 - np.exp(-t / 1000 / time_const))

            self.time_series_data.current.append(current_value)

            if i == 0:
                didt_value = 0
            elif i == 1:
                didt_value = self.inputs.sourceVoltage / inductance_to_fault
            else:
                didt_value = (current_value - prev_current) / dt

            self.time_series_data.diDt.append(didt_value / 1000)

            if abs(didt_value) / 1000 > self.time_series_data.peakDiDt:
                self.time_series_data.peakDiDt = abs(didt_value) / 1000

            if abs(current_value) > self.time_series_data.peakCurrent:
                self.time_series_data.peakCurrent = abs(current_value)

            voltage_value = inductance_to_fault * didt_value
            self.time_series_data.voltage.append(voltage_value)

            prev_current = current_value

        self.apply_low_pass_filter(100)
        self.results.peakDiDt = self.time_series_data.peakDiDt

        if num_points > 0:
            self.results.steadyStateDiDt = self.time_series_data.diDt[-1]

    def apply_low_pass_filter(self, cutoff_freq):
        """Apply 2nd order low-pass filter to di/dt values"""
        dt = (self.time_series_data.timePoints[1] - self.time_series_data.timePoints[0]) / 1000
        omega = 2 * np.pi * cutoff_freq
        c = 1 / np.tan(omega * dt / 2)

        a0 = c ** 2 + np.sqrt(2) * c + 1
        a1 = 2 * (1 - c ** 2)
        a2 = c ** 2 - np.sqrt(2) * c + 1

        b0 = 1
        b1 = 2
        b2 = 1

        b0 /= a0
        b1 /= a0
        b2 /= a0
        a1 /= a0
        a2 /= a0

        x = self.time_series_data.diDt
        y = [0] * len(x)

        y[0] = b0 * x[0]
        if len(x) > 1:
            y[1] = b0 * x[1] + b1 * x[0] - a1 * y[0]

        for i in range(2, len(x)):
            y[i] = b0 * x[i] + b1 * x[i - 1] + b2 * x[i - 2] - a1 * y[i - 1] - a2 * y[i - 2]

        self.time_series_data.filteredDiDt = y

    def create_excel_report(self):
        """Create Excel report and return as bytes"""
        # Create input parameters dataframe
        inputs_data = {
            'Parameter': [
                'Source Voltage (VDC)',
                'Source Internal Resistance (Ohms)',
                'Line Resistance (Ohms/km)',
                'Fault Resistance (Ohms)',
                'Fault Distance from Nearest Substation (km)',
                'Number of Substations',
                'Substation Spacing (km)',
                'Protection Trip Time (seconds)',
                'Protection Trip Current (Amps)',
                'System Inductance (H)',
                'Line Inductance (H/km)'
            ],
            'Value': [
                self.inputs.sourceVoltage,
                self.inputs.sourceInternalResistance,
                self.inputs.lineResistance,
                self.inputs.faultResistance,
                self.inputs.faultDistance,
                self.inputs.numSubstations,
                self.inputs.substationSpacing,
                self.inputs.protectionTripTime,
                self.inputs.protectionTripCurrent,
                self.inputs.systemInductance,
                self.inputs.lineInductance
            ]
        }

        results_data = {
            'Result': [
                'Total Fault Current (Amps)',
                'Energy Released (Joules)',
                'Thermal Stress (A*sqrt(s))',
                'Protection Will Trip',
                'Trip Time (s)',
                'Shortest Distance Fault Current (Amps)',
                'Longest Distance Fault Current (Amps)',
                'Shortest Distance Current Rate of Change (kA/s)',
                'Longest Distance Current Rate of Change (kA/s)',
                'Peak Current Rate of Change (kA/s)',
                'Steady State Current Rate of Change (kA/s)'
            ],
            'Value': [
                self.results.totalFaultCurrent,
                self.results.energyReleased,
                self.results.thermalStress,
                "Yes" if self.results.protectionWillTrip else "No",
                self.results.tripTime,
                self.results.shortestDistanceCurrent,
                self.results.longestDistanceCurrent,
                self.results.shortestDistanceDiDt,
                self.results.longestDistanceDiDt,
                self.results.peakDiDt,
                self.results.steadyStateDiDt
            ]
        }

        substation_data = {
            'Substation': [f"Substation {i + 1}" for i in range(self.inputs.numSubstations)],
            'Location (km)': self.substation_locations,
            'Distance to Fault (km)': [abs(self.inputs.faultDistance - loc) for loc in self.substation_locations],
            'Fault Current (Amps)': self.fault_currents,
            'Voltage Drop (V)': self.voltage_drops,
            'Contribution (%)': [(fc / self.results.totalFaultCurrent * 100) if self.results.totalFaultCurrent > 0 else 0
                           for fc in self.fault_currents]
        }

        df_inputs = pd.DataFrame(inputs_data)
        df_results = pd.DataFrame(results_data)
        df_substations = pd.DataFrame(substation_data)

        # Create time domain data if available
        time_data = None
        if self.time_series_data.timePoints:
            time_data = {
                'Time (ms)': self.time_series_data.timePoints,
                'Electric Current (A)': self.time_series_data.current,
                'di/dt (kA/s)': self.time_series_data.diDt,
                'Filtered di/dt (kA/s)': self.time_series_data.filteredDiDt,
                'Voltage (V)': self.time_series_data.voltage
            }
            df_time = pd.DataFrame(time_data)

        # Create Excel file in memory
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df_inputs.to_excel(writer, sheet_name='Input', index=False)
            df_results.to_excel(writer, sheet_name='Results', index=False)
            df_substations.to_excel(writer, sheet_name='Generator', index=False)
            
            if time_data:
                df_time.to_excel(writer, sheet_name='Time Domain Data', index=False)

            # Apply formatting
            workbook = writer.book
            from openpyxl.styles import PatternFill, Border, Side, Font
            from openpyxl.utils import get_column_letter

            blue_fill = PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                top=Side(style='thin'), bottom=Side(style='thin'))
            consolas_font = Font(name='Consolas')

            for sheet_name in ['Input', 'Results', 'Generator']:
                if sheet_name in workbook.sheetnames:
                    worksheet = workbook[sheet_name]
                    if sheet_name == 'Input':
                        df = df_inputs
                    elif sheet_name == 'Results':
                        df = df_results
                    else:
                        df = df_substations
                    
                    num_rows = len(df)
                    num_cols = len(df.columns)

                    for row in range(1, num_rows + 2):
                        for col in range(1, num_cols + 1):
                            cell = worksheet.cell(row=row, column=col)
                            cell.border = thin_border
                            cell.font = consolas_font
                            if col == 1 and row > 1 and sheet_name in ['Input', 'Results']:
                                cell.fill = blue_fill

                    for col in range(1, num_cols + 1):
                        column_letter = get_column_letter(col)
                        max_length = 0
                        for row in range(1, num_rows + 2):
                            cell_value = worksheet.cell(row=row, column=col).value
                            if cell_value:
                                max_length = max(max_length, len(str(cell_value)))
                        adjusted_width = max_length + 4
                        worksheet.column_dimensions[column_letter].width = adjusted_width

        buffer.seek(0)
        return buffer.getvalue()


def simulation_home(request):
    """Main simulation page with input form"""
    form = SimulationForm()
    return render(request, 'simulation/home.html', {'form': form})


def run_simulation(request):
    """Run the simulation calculation"""
    if request.method == 'POST':
        form = SimulationForm(request.POST)
        if form.is_valid():
            # Save the simulation run
            simulation_run = form.save(commit=False)
            if request.user.is_authenticated:
                simulation_run.user = request.user
            
            # Create inputs object from form data
            inputs = SimulationInputs(
                sourceVoltage=form.cleaned_data['source_voltage'],
                sourceInternalResistance=form.cleaned_data['source_internal_resistance'],
                lineResistance=form.cleaned_data['line_resistance'],
                faultResistance=form.cleaned_data['fault_resistance'],
                faultDistance=form.cleaned_data['fault_distance'],
                numSubstations=form.cleaned_data['num_substations'],
                substationSpacing=form.cleaned_data['substation_spacing'],
                protectionTripTime=form.cleaned_data['protection_trip_time'],
                protectionTripCurrent=form.cleaned_data['protection_trip_current'],
                systemInductance=form.cleaned_data['system_inductance'],
                lineInductance=form.cleaned_data['line_inductance'],
                simulationTimeStep=form.cleaned_data['simulation_time_step'],
                simulationDuration=form.cleaned_data['simulation_duration'],
                lineLength=form.cleaned_data['line_length']
            )
            
            try:
                # Run calculation
                calculator = DCShortCircuitCalculation(inputs)
                calculator.calculate_short_circuit()
                
                # Update simulation run with results
                simulation_run.total_fault_current = calculator.results.totalFaultCurrent
                simulation_run.energy_released = calculator.results.energyReleased
                simulation_run.thermal_stress = calculator.results.thermalStress
                simulation_run.protection_will_trip = calculator.results.protectionWillTrip
                simulation_run.trip_time = calculator.results.tripTime
                simulation_run.shortest_distance_current = calculator.results.shortestDistanceCurrent
                simulation_run.longest_distance_current = calculator.results.longestDistanceCurrent
                simulation_run.shortest_distance_didt = calculator.results.shortestDistanceDiDt
                simulation_run.longest_distance_didt = calculator.results.longestDistanceDiDt
                simulation_run.peak_didt = calculator.results.peakDiDt
                simulation_run.steady_state_didt = calculator.results.steadyStateDiDt
                
                # Generate Excel report
                excel_data = calculator.create_excel_report()
                excel_file = ContentFile(excel_data, name=f'simulation_{simulation_run.id}.xlsx')
                simulation_run.excel_file.save(f'simulation_{simulation_run.id}.xlsx', excel_file)
                
                # Save the simulation run
                simulation_run.save()
                
                # Store calculator in session for chart generation
                request.session[f'calculator_{simulation_run.id}'] = {
                    'substation_locations': calculator.substation_locations,
                    'fault_currents': calculator.fault_currents,
                    'voltage_drops': calculator.voltage_drops,
                    'time_series_data': {
                        'timePoints': calculator.time_series_data.timePoints,
                        'current': calculator.time_series_data.current,
                        'diDt': calculator.time_series_data.diDt,
                        'filteredDiDt': calculator.time_series_data.filteredDiDt,
                        'voltage': calculator.time_series_data.voltage
                    },
                    'results': {
                        'shortestDistanceCurrent': calculator.results.shortestDistanceCurrent,
                        'longestDistanceCurrent': calculator.results.longestDistanceCurrent,
                        'shortestDistanceDiDt': calculator.results.shortestDistanceDiDt,
                        'longestDistanceDiDt': calculator.results.longestDistanceDiDt,
                    }
                }
                
                messages.success(request, 'Simulation completed successfully!')
                return redirect('simulation:results', simulation_id=simulation_run.id)
                
            except Exception as e:
                messages.error(request, f'Error running simulation: {str(e)}')
                return render(request, 'simulation/home.html', {'form': form})
        else:
            messages.error(request, 'Please correct the errors in the form.')
            return render(request, 'simulation/home.html', {'form': form})
    
    return redirect('simulation:home')


def simulation_results(request, simulation_id):
    """Display simulation results"""
    simulation_run = get_object_or_404(SimulationRun, id=simulation_id)
    
    context = {
        'simulation': simulation_run,
        'has_charts': f'calculator_{simulation_id}' in request.session
    }
    
    return render(request, 'simulation/results.html', context)


def download_excel(request, simulation_id):
    """Download Excel report"""
    simulation_run = get_object_or_404(SimulationRun, id=simulation_id)
    
    if simulation_run.excel_file:
        response = HttpResponse(
            simulation_run.excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="simulation_{simulation_id}.xlsx"'
        return response
    else:
        raise Http404("Excel file not found")


import os
from django.conf import settings
import uuid


def get_chart(request, simulation_id, chart_type):
    """生成并返回图表文件URL"""
    # 导入所需库
    import matplotlib
    matplotlib.use('Agg')  # 设置非交互式后端
    import matplotlib.pyplot as plt
    import numpy as np
    import time
    import os
    import random
    from django.conf import settings
    from django.http import JsonResponse

    # 打印调试信息
    print(f"开始生成图表: {chart_type}, simulation_id: {simulation_id}")

    session_key = f'calculator_{simulation_id}'

    # 重置matplotlib状态
    plt.close('all')

    if session_key not in request.session:
        print(f"会话中找不到数据: {session_key}")
        return JsonResponse({'error': '图表数据不可用'}, status=404)

    data = request.session[session_key]
    fig = None

    try:
        # 确定是否为时间序列图表
        is_time_series = chart_type in ['current_vs_time', 'didt_vs_time', 'voltage_vs_time']

        # 创建保存目录
        charts_dir = os.path.join(settings.MEDIA_ROOT, 'charts')
        os.makedirs(charts_dir, exist_ok=True)

        # 生成唯一文件名
        timestamp = int(time.time())
        random_str = random.randint(10000, 99999)
        filename = f"{chart_type}_{simulation_id}_{timestamp}_{random_str}.png"
        filepath = os.path.join(charts_dir, filename)

        print(f"数据结构: {list(data.keys())}")

        # 创建图表
        fig = plt.figure(figsize=(12, 7), dpi=100, facecolor='white')
        ax = fig.add_subplot(111)
        ax.set_facecolor('#f8f8f8')

        # 根据图表类型绘制不同图表
        if chart_type == 'fault_currents':
            fault_currents = data['fault_currents']
            print(f"故障电流数据: {fault_currents}")

            num_substations = len(fault_currents)
            bars = ax.bar(range(1, num_substations + 1), fault_currents, color='blue')

            # 添加数值标签
            for bar in bars:
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width() / 2., height + 5,
                        f'{height:.1f}', ha='center', va='bottom')

            ax.set_title('Fault Current Contribution by Substation', fontsize=14, fontweight='bold')
            ax.set_xlabel('Substation', fontsize=12)
            ax.set_ylabel('Current (Amps)', fontsize=12)
            ax.set_xticks(range(1, num_substations + 1))
            ax.grid(True, linestyle='--', alpha=0.7)

        elif chart_type == 'voltage_drops':
            voltage_drops = data['voltage_drops']
            print(f"电压降数据: {voltage_drops}")

            num_substations = len(voltage_drops)
            bars = ax.bar(range(1, num_substations + 1), voltage_drops, color='orange')

            # 添加数值标签
            for bar in bars:
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width() / 2., height + 0.1,
                        f'{height:.2f}', ha='center', va='bottom')

            ax.set_title('Voltage Drop by Substation', fontsize=14, fontweight='bold')
            ax.set_xlabel('Substation', fontsize=12)
            ax.set_ylabel('Voltage Drop (V)', fontsize=12)
            ax.set_xticks(range(1, num_substations + 1))
            ax.grid(True, linestyle='--', alpha=0.7)

        elif chart_type == 'critical_faults':
            locations = ['Shortest Distance', 'Longest Distance']
            currents = [data['results']['shortestDistanceCurrent'],
                        data['results']['longestDistanceCurrent']]
            didts = [data['results']['shortestDistanceDiDt'],
                     data['results']['longestDistanceDiDt']]

            print(f"临界故障数据 - 电流: {currents}, di/dt: {didts}")

            x = np.arange(len(locations))
            width = 0.35

            # 主Y轴
            color1 = 'tab:blue'
            ax.set_xlabel('Location', fontsize=12)
            ax.set_ylabel('Fault Current (A)', color=color1, fontsize=12)
            bars1 = ax.bar(x - width / 2, currents, width, label='Fault Current (A)', color=color1)
            ax.tick_params(axis='y', labelcolor=color1)

            # 为电流添加数值标签
            for bar in bars1:
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width() / 2., height + 5,
                        f'{height:.1f}', ha='center', va='bottom', color=color1)

            # 次Y轴
            ax2 = ax.twinx()
            color2 = 'tab:red'
            ax2.set_ylabel('Current Rate of Change (kA/s)', color=color2, fontsize=12)
            bars2 = ax2.bar(x + width / 2, didts, width, label='Current Rate of Change (kA/s)', color=color2)
            ax2.tick_params(axis='y', labelcolor=color2)

            # 为di/dt添加数值标签
            for bar in bars2:
                height = bar.get_height()
                ax2.text(bar.get_x() + bar.get_width() / 2., height + 5,
                         f'{height:.1f}', ha='center', va='bottom', color=color2)

            ax.set_title('Critical Fault Analysis', fontsize=14, fontweight='bold')
            ax.set_xticks(x)
            ax.set_xticklabels(locations)

            # 添加图例
            fig.legend(loc='upper center', bbox_to_anchor=(0.5, 0.05), ncol=2)

        elif chart_type == 'current_vs_time':
            time_data = data['time_series_data']
            timePoints = time_data['timePoints']
            current = time_data['current']

            print(f"电流时间序列数据点数: {len(timePoints)}")
            print(f"电流值范围: {min(current) if current else 'N/A'} - {max(current) if current else 'N/A'}")

            if not timePoints or not current or len(timePoints) != len(current):
                return JsonResponse({'error': '时间序列数据无效'}, status=400)

            # 降采样以减小文件大小
            if len(timePoints) > 500:
                step = len(timePoints) // 500
                timePoints = timePoints[::step]
                current = current[::step]
                print(f"降采样后数据点数: {len(timePoints)}")

            ax.plot(timePoints, current, label='Fault Current', color='green', linewidth=2.5)
            ax.set_title('Fault Current vs Time', fontsize=14, fontweight='bold')
            ax.set_xlabel('Time (ms)', fontsize=12)
            ax.set_ylabel('Current (A)', fontsize=12)
            ax.grid(True, linestyle='--', alpha=0.7)
            ax.legend()

        elif chart_type == 'didt_vs_time':
            time_data = data['time_series_data']
            timePoints = time_data['timePoints']
            diDt = time_data['diDt']
            filteredDiDt = time_data.get('filteredDiDt', [])

            print(f"di/dt时间序列数据点数: {len(timePoints)}")
            print(f"di/dt值范围: {min(diDt) if diDt else 'N/A'} - {max(diDt) if diDt else 'N/A'}")

            if not timePoints or not diDt or len(timePoints) != len(diDt):
                return JsonResponse({'error': 'di/dt时间序列数据无效'}, status=400)

            # 降采样以减小文件大小
            if len(timePoints) > 500:
                step = len(timePoints) // 500
                timePoints = timePoints[::step]
                diDt = diDt[::step]
                if filteredDiDt and len(filteredDiDt) == len(time_data['timePoints']):
                    filteredDiDt = filteredDiDt[::step]
                print(f"降采样后数据点数: {len(timePoints)}")

            ax.plot(timePoints, diDt, label='Raw Data', color='blue', linewidth=1.0)
            if filteredDiDt and len(filteredDiDt) == len(timePoints):
                ax.plot(timePoints, filteredDiDt, label='Low-Pass Filtered Data (100Hz)',
                        color='red', linewidth=2.0)
            ax.set_title('Current Rate of Change (di/dt) vs Time', fontsize=14, fontweight='bold')
            ax.set_xlabel('Time (ms)', fontsize=12)
            ax.set_ylabel('di/dt (kA/s)', fontsize=12)
            ax.grid(True, linestyle='--', alpha=0.7)
            ax.legend()

        elif chart_type == 'voltage_vs_time':
            time_data = data['time_series_data']
            timePoints = time_data['timePoints']
            voltage = time_data['voltage']

            print(f"电压时间序列数据点数: {len(timePoints)}")
            print(f"电压值范围: {min(voltage) if voltage else 'N/A'} - {max(voltage) if voltage else 'N/A'}")

            if not timePoints or not voltage or len(timePoints) != len(voltage):
                return JsonResponse({'error': '电压时间序列数据无效'}, status=400)

            # 降采样以减小文件大小
            if len(timePoints) > 500:
                step = len(timePoints) // 500
                timePoints = timePoints[::step]
                voltage = voltage[::step]
                print(f"降采样后数据点数: {len(timePoints)}")

            ax.plot(timePoints, voltage, label='Voltage', color='purple', linewidth=2.5)
            ax.set_title('Voltage vs Time', fontsize=14, fontweight='bold')
            ax.set_xlabel('Time (ms)', fontsize=12)
            ax.set_ylabel('Voltage (V)', fontsize=12)
            ax.grid(True, linestyle='--', alpha=0.7)
            ax.legend()

        else:
            return JsonResponse({'error': f'无效的图表类型: {chart_type}'}, status=400)

        # 完成图表布局设置
        fig.tight_layout(pad=3.0)

        # 确保图表完全渲染
        fig.canvas.draw()

        # 为时间序列图表使用较低DPI减小文件大小
        dpi_value = 100 if is_time_series else 150

        # 保存图表
        print(f"保存图表到: {filepath}")
        fig.savefig(filepath, format='png', dpi=dpi_value, bbox_inches='tight',
                    facecolor=fig.get_facecolor(), transparent=False)

        # 验证文件保存成功
        if os.path.exists(filepath):
            filesize = os.path.getsize(filepath)
            print(f"图表已保存，文件大小: {filesize} 字节")
            if filesize < 1000:
                print("警告: 图表文件过小，可能不完整")
                return JsonResponse({'error': '图表文件过小'}, status=500)
        else:
            print(f"错误: 文件未能保存到 {filepath}")
            return JsonResponse({'error': '保存图表失败'}, status=500)

        # 构建完整URL
        file_url = f"{settings.MEDIA_URL}charts/{filename}"

        # 返回图表URL
        print(f"返回图表URL: {file_url}")
        return JsonResponse({'image_url': file_url})

    except Exception as e:
        import traceback
        print(f"生成图表错误: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)
    finally:
        # 确保资源释放
        if fig is not None:
            plt.close(fig)
        plt.close('all')  # 额外安全措施


def simulation_history(request):
    """Display history of simulations"""
    # 获取所有模拟记录，按创建时间降序排列
    simulations = SimulationRun.objects.all().order_by('-created_at')

    context = {
        'simulations': simulations
    }

    return render(request, 'simulation/history.html', context)

